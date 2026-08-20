"""Build the combined UAT feedback workbook (one file, a Findings tab per tester).

v2 (2026-08-19): the steps are PARSED from the plan markdown — the same source
build-artifact.py uses — so the plan, the published page, and this workbook
cannot drift. v1 kept its own hardcoded copy of all 240 step labels; that copy
is gone.

The output xlsx is uploaded to Google Drive with conversion to a Google Sheet
(tabs: README, Julia Findings, Nate Findings, Decisions). The repo copy is the
blank template; live answers accumulate in the Sheet, not here.

Usage: python build-feedback-workbook.py <plan.md> <out.xlsx>
"""
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Arial"

# ------------------------------------------------------------------ md parsing
def plain(s):
    """Markdown inline → plain text for a cell."""
    s = s.strip()
    s = re.sub(r"\*\*(.+?)\*\*", r"\1", s)
    s = re.sub(r"`(.+?)`", r"\1", s)
    s = re.sub(r"(?<!\*)\*([^*]+)\*(?!\*)", r"\1", s)
    return s


def parse_lanes(md):
    lane_re = re.compile(r"^## Lane ([A-Z]) — (.+?)$", re.M)
    lanes = []
    matches = list(lane_re.finditer(md))
    for i, m in enumerate(matches):
        letter, name = m.group(1), m.group(2).replace("⭐", "").strip()
        end = matches[i + 1].start() if i + 1 < len(matches) else md.find("\n# Appendix A")
        body = md[m.end():end]
        steps = []
        for line in body.splitlines():
            line = line.strip()
            if not line.startswith("|"):
                continue
            cells = [c.strip() for c in line.strip("|").split(" | ")]
            if len(cells) != 3:
                continue
            sid = cells[0].strip()
            if not re.fullmatch(r"[A-Z]-\d{2}", sid):
                continue
            steps.append((sid, plain(cells[1])))
        lanes.append((letter, name, steps))
    return lanes


DECISIONS = [
    ("D-1", "Camps and club managers",
     "Spec SS-G says camps are individual self-registration only, but a club manager can still "
     "register athletes for one. Block it outright, or keep it as a convenience?"),
    ("D-2", "Host payout timing",
     "The 'owed' formula is settled (gross before fees, refunds not deducted). WHEN payout "
     "happens is still just wording on the host page. Is '1 week after the event' the policy?"),
    ("D-3", "Hosts and SMS",
     "Hosts currently get email only from the event Communicate page; SMS stays league-admin-only "
     "because each text bills UCG. Still right?"),
    ("D-4", "Invoice numbering format",
     "Two formats coexist. Which one do you want on real financial records?"),
    ("D-5", "Refund policy edges",
     "After-deadline is 75% before processing fees. Should add-ons follow the same 75% rule, "
     "or refund in full?"),
    ("D-6", "Anything in the known-gaps appendix that should not wait",
     "List anything from Appendix A you want prioritized, and why."),
]

RESULTS = ["PASS", "FAIL", "MISSING", "UNCLEAR", "BLOCKED", "N/A"]
SEVERITIES = ["S1", "S2", "S3", "S4", "Q", "D"]
YESNO = ["Y", "N", "N/A"]

HEADERS = [
    ("Step ID", 9), ("Lane", 7), ("Step (what to do)", 56),
    ("Result", 11), ("Finding ID", 12), ("Severity", 10),
    ("What you expected", 40), ("What actually happened (exact error text)", 48),
    ("Timestamp (local + tz)", 21), ("Hard reload fix it?", 17),
    ("Device / browser", 20), ("Signed in as", 18),
    ("Screenshot(s)", 26), ("Reported in-app?", 15), ("Notes", 34),
]

NAVY, CORAL, YELLOW, GREY = "1F3352", "F0785A", "FFF6CC", "F4F4F4"
thin = Side(style="thin", color="C6CDD6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30


def build_findings_tab(wb, tab_name, initials, lanes):
    fs = wb.create_sheet(tab_name)
    for i, (title, width) in enumerate(HEADERS, start=1):
        fs.cell(row=1, column=i, value=title)
        fs.column_dimensions[get_column_letter(i)].width = width
    style_header(fs, 1, len(HEADERS))
    fs.freeze_panes = "D2"

    example = [
        "M-07", None, "EXAMPLE ROW - delete or overwrite me",
        "FAIL", "M-07-01", "S1",
        "3-D Secure challenge appears, then the payment completes",
        'Challenge appeared, I approved it, then "Something went wrong" and the cart still had the item',
        "2026-08-19 14:32 PT", "N", "Win 11 / Chrome 141", "Athlete (ZZTEST-Robin)",
        f"M-07-01_01_{initials}.png", "Y",
        "Reproduced twice. Stripe dashboard shows no payment attempt at all.",
    ]
    for i, val in enumerate(example, start=1):
        c = fs.cell(row=2, column=i, value=val)
        c.font = Font(name=FONT, size=10, italic=True, color="8A7A2E")
        c.fill = PatternFill("solid", fgColor=YELLOW)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = BORDER
    fs.cell(row=2, column=2, value="M")
    fs.row_dimensions[2].height = 42

    row = 3
    for letter, lane_name, steps in lanes:
        fs.cell(row=row, column=1, value=f"{letter}  -  {lane_name}")
        fs.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        b = fs.cell(row=row, column=1)
        b.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
        b.fill = PatternFill("solid", fgColor=CORAL)
        b.alignment = Alignment(vertical="center", indent=1)
        fs.row_dimensions[row].height = 20
        row += 1
        for sid, label in steps:
            a = fs.cell(row=row, column=1, value=sid)
            a.font = Font(name=FONT, size=10, bold=True, color=NAVY)
            a.fill = PatternFill("solid", fgColor=GREY)
            b = fs.cell(row=row, column=2, value=letter)
            b.font = Font(name=FONT, size=10)
            b.fill = PatternFill("solid", fgColor=GREY)
            cc = fs.cell(row=row, column=3, value=label)
            cc.font = Font(name=FONT, size=10)
            cc.fill = PatternFill("solid", fgColor=GREY)
            cc.alignment = Alignment(vertical="top", wrap_text=True)
            d = fs.cell(row=row, column=4)
            d.fill = PatternFill("solid", fgColor=YELLOW)
            row += 1

    fs.cell(row=row, column=1, value="EXTRA  -  findings not tied to a step (use e.g. M-X-01)")
    fs.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    b = fs.cell(row=row, column=1)
    b.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    b.fill = PatternFill("solid", fgColor=NAVY)
    b.alignment = Alignment(vertical="center", indent=1)
    row += 1
    for _ in range(40):
        fs.cell(row=row, column=1).fill = PatternFill("solid", fgColor=YELLOW)
        fs.cell(row=row, column=4).fill = PatternFill("solid", fgColor=YELLOW)
        row += 1
    last_row = row - 1

    dv_result = DataValidation(type="list", formula1=f'"{",".join(RESULTS)}"', allow_blank=True)
    dv_sev = DataValidation(type="list", formula1=f'"{",".join(SEVERITIES)}"', allow_blank=True)
    dv_yn = DataValidation(type="list", formula1=f'"{",".join(YESNO)}"', allow_blank=True)
    dv_inapp = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    for dv in (dv_result, dv_sev, dv_yn, dv_inapp):
        fs.add_data_validation(dv)
    dv_result.add(f"D2:D{last_row}")
    dv_sev.add(f"F2:F{last_row}")
    dv_yn.add(f"J2:J{last_row}")
    dv_inapp.add(f"N2:N{last_row}")
    fs.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"


def build(md_path, out_path):
    md = open(md_path, encoding="utf-8").read()
    lanes = parse_lanes(md)
    total = sum(len(s) for _, _, s in lanes)
    assert total == 240, f"expected 240 steps, parsed {total}"

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False

    rows = [
        ("UCG Preflight - Feedback Tracker", "title"),
        ("One shared sheet. Julia works in 'Julia Findings', Nate in 'Nate Findings'.", "sub"),
        ("Companion to the UCG Preflight page / docs/plans/2026-08-19-uat-stress-test-plan.md", "sub"),
        ("", None),
        ("HOW TO USE", "h"),
        ("1. Work in YOUR OWN Findings tab. Every step from the plan is already there, in order.", "p"),
        ("2. Fill in the Result column for EVERY step - including the ones that pass.", "p"),
        ("   A PASS row takes two seconds and is what makes the untested gaps visible.", "p"),
        ("3. Only fill the rest of a row when it is NOT a plain PASS.", "p"),
        ("4. Anything BROKEN: file it in-app first (Report a problem), THEN add the row here.", "p"),
        ("   The in-app widget captures the console-error buffer, which is overwritten as you", "p"),
        ("   keep clicking. Start that description with the Finding ID and your initials.", "p"),
        ("5. Blank rows at the bottom of each Findings tab are for anything not tied to a step.", "p"),
        ("6. Julia: the Decisions tab needs no testing - just answers.", "p"),
        ("", None),
        ("BEFORE YOU FILE 'IT DIDN'T UPDATE'", "h"),
        ("Hard-reload once (Ctrl+Shift+R / Cmd+Shift+R), then answer the 'Hard reload fix it?'", "p"),
        ("column. Live updates are wired for SCORES ONLY - most other cross-device changes", "p"),
        ("legitimately need a refresh. This one column saves a triage cycle every time.", "p"),
        ("", None),
        ("TEST DATA", "h"),
        ("Prefix everything you create with  ZZTEST-  so the pre-launch sweep is a search,", "p"),
        ("not an excavation. Forgot? Say so in Notes. Do NOT delete as you go - a broken row", "p"),
        ("is often the evidence.", "p"),
        ("", None),
        ("SCREENSHOTS", "h"),
        ("Filename:  <Finding ID>_<seq>_<your initials>.png     e.g.  M-07-01_01_JB.png", "p"),
        ("Include the address bar. Annotate the exact element. Use a screen RECORDING for", "p"),
        ("anything with motion. For money bugs, capture line items AND the total in one frame.", "p"),
        ("", None),
        ("SEVERITY", "h"),
        ("S1  Blocker - a core flow can't complete, or money/data is wrong", "p"),
        ("S2  Major - broken or badly wrong, but there is a workaround", "p"),
        ("S3  Minor - works, but wrong in a way a user would notice", "p"),
        ("S4  Polish - cosmetic", "p"),
        ("Q   Question - you can't tell whether it's wrong", "p"),
        ("D   Decision needed - works as built, but is it what we want?  <- most valuable", "p"),
        ("", None),
        ("RESULT CODES", "h"),
        ("PASS / FAIL / MISSING (described behavior doesn't exist) / UNCLEAR (couldn't tell) /", "p"),
        ("BLOCKED (couldn't get to it) / N/A (doesn't apply to you)", "p"),
        ("", None),
        ("PROGRESS", "h"),
    ]
    r = 1
    for text, kind in rows:
        c = ws.cell(row=r, column=1, value=text)
        if kind == "title":
            c.font = Font(name=FONT, bold=True, size=16, color=NAVY)
            ws.row_dimensions[r].height = 22
        elif kind == "sub":
            c.font = Font(name=FONT, size=10, italic=True, color="5A6875")
        elif kind == "h":
            c.font = Font(name=FONT, bold=True, size=11, color=NAVY)
        else:
            c.font = Font(name=FONT, size=10)
        r += 1

    # Two-column progress block. Ranges start at row 3 (example row never counts);
    # "?-??" matches step ids (A-01) but not lane banners or extra ids (M-X-01).
    ws.cell(row=r, column=2, value="Julia").font = Font(name=FONT, bold=True, size=10, color=NAVY)
    ws.cell(row=r, column=3, value="Nate").font = Font(name=FONT, bold=True, size=10, color=NAVY)
    r += 1
    def refs(tpl):
        return (tpl.format(tab="'Julia Findings'"), tpl.format(tab="'Nate Findings'"))
    counters = [
        ("Steps in the plan", '=COUNTIF({tab}!A3:A2000,"?-??")'),
        ("Steps with a Result filled in", '=COUNTIFS({tab}!A3:A2000,"?-??",{tab}!D3:D2000,"<>")'),
        ("Still to do", '=COUNTIF({tab}!A3:A2000,"?-??")-COUNTIFS({tab}!A3:A2000,"?-??",{tab}!D3:D2000,"<>")'),
        ("", None),
        ("PASS", '=COUNTIF({tab}!D3:D2000,"PASS")'),
        ("FAIL", '=COUNTIF({tab}!D3:D2000,"FAIL")'),
        ("MISSING", '=COUNTIF({tab}!D3:D2000,"MISSING")'),
        ("UNCLEAR", '=COUNTIF({tab}!D3:D2000,"UNCLEAR")'),
        ("BLOCKED", '=COUNTIF({tab}!D3:D2000,"BLOCKED")'),
        ("", None),
        ("S1 blockers", '=COUNTIF({tab}!F3:F2000,"S1")'),
        ("S2 major", '=COUNTIF({tab}!F3:F2000,"S2")'),
        ("Decisions raised (D)", '=COUNTIF({tab}!F3:F2000,"D")'),
    ]
    for label, tpl in counters:
        if tpl is None:
            r += 1
            continue
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10)
        j, n = refs(tpl)
        for col, f in ((2, j), (3, n)):
            cell = ws.cell(row=r, column=col, value=f)
            cell.font = Font(name=FONT, size=10, bold=True, color=NAVY)
            cell.alignment = Alignment(horizontal="left")
        r += 1
    ws.column_dimensions["A"].width = 92
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10

    build_findings_tab(wb, "Julia Findings", "JS", lanes)
    build_findings_tab(wb, "Nate Findings", "NS", lanes)

    ds = wb.create_sheet("Decisions")
    ds.sheet_view.showGridLines = False
    ds.cell(row=1, column=1, value="Decisions needed - no testing required, just Julia's answer")
    ds.cell(row=1, column=1).font = Font(name=FONT, bold=True, size=14, color=NAVY)
    ds.merge_cells("A1:D1")
    for i, h in enumerate(["#", "Topic", "The question", "Julia's answer"], start=1):
        ds.cell(row=3, column=i, value=h)
    style_header(ds, 3, 4)
    for i, (num, topic, q) in enumerate(DECISIONS, start=4):
        ds.cell(row=i, column=1, value=num).font = Font(name=FONT, bold=True, size=10, color=NAVY)
        ds.cell(row=i, column=2, value=topic).font = Font(name=FONT, size=10, bold=True)
        ds.cell(row=i, column=3, value=q).font = Font(name=FONT, size=10)
        ans = ds.cell(row=i, column=4)
        ans.fill = PatternFill("solid", fgColor=YELLOW)
        ans.font = Font(name=FONT, size=10)
        for col in range(1, 5):
            cc = ds.cell(row=i, column=col)
            cc.alignment = Alignment(vertical="top", wrap_text=True)
            cc.border = BORDER
        ds.row_dimensions[i].height = 58
    for col, w in zip("ABCD", (6, 30, 66, 52)):
        ds.column_dimensions[col].width = w

    wb.save(out_path)
    print(f"wrote {out_path}: {total} steps x 2 tabs")


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2])

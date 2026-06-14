"""Build Supabase import from ScoreFlippers exports.

Inputs (same dir):
  clubs_raw.json     - scraped club profiles (222), keyed by ScoreFlippers ClubID
  athletes_raw.xlsx  - League athletes export (2617)

Outputs (same dir):
  clubs.json    - app Club shape (reference)
  people.json   - app Athlete shape (reference)
  import.sql    - idempotent load: wipe demo/transactional data, insert real
                  clubs + people. Run in the Supabase SQL editor (bypasses RLS).
  REPORT.md     - mapping summary + flagged ambiguities
"""
import json, os, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
def p(n): return os.path.join(HERE, n)

# ---------------------------------------------------------------- region/state
ENUM_REGIONS = {'Northeast','Mid-Atlantic','Southeast','Mideast','Midwest',
                'South Central','West','Other'}
STATE_REGIONS = {
 'Alabama':'Southeast','Alaska':'West','Arizona':'West','Arkansas':'South Central',
 'California':'West','Colorado':'West','Connecticut':'Northeast','Delaware':'Mid-Atlantic',
 'District of Columbia':'Mid-Atlantic','Florida':'Southeast','Georgia':'Southeast',
 'Hawaii':'West','Idaho':'West','Illinois':'Midwest','Indiana':'Mideast','Iowa':'Midwest',
 'Kansas':'South Central','Kentucky':'Mideast','Louisiana':'Southeast','Maine':'Northeast',
 'Maryland':'Mid-Atlantic','Massachusetts':'Northeast','Michigan':'Mideast',
 'Minnesota':'Midwest','Mississippi':'Southeast','Missouri':'Midwest','Montana':'West',
 'Nebraska':'Midwest','Nevada':'West','New Hampshire':'Northeast','New Jersey':'Mid-Atlantic',
 'New Mexico':'West','New York':'Northeast','North Carolina':'Southeast',
 'North Dakota':'Midwest','Ohio':'Mideast','Oklahoma':'South Central','Oregon':'West',
 'Pennsylvania':'Mid-Atlantic','Rhode Island':'Northeast','South Carolina':'Southeast',
 'South Dakota':'Midwest','Tennessee':'Southeast','Texas':'South Central','Utah':'West',
 'Vermont':'Northeast','Virginia':'Mid-Atlantic','Washington':'West',
 'West Virginia':'Mid-Atlantic','Wisconsin':'Midwest','Wyoming':'West',
}
ABBR_FULL = {
 'AL':'Alabama','AK':'Alaska','AZ':'Arizona','AR':'Arkansas','CA':'California',
 'CO':'Colorado','CT':'Connecticut','DE':'Delaware','DC':'District of Columbia',
 'FL':'Florida','GA':'Georgia','HI':'Hawaii','ID':'Idaho','IL':'Illinois','IN':'Indiana',
 'IA':'Iowa','KS':'Kansas','KY':'Kentucky','LA':'Louisiana','ME':'Maine','MD':'Maryland',
 'MA':'Massachusetts','MI':'Michigan','MN':'Minnesota','MS':'Mississippi','MO':'Missouri',
 'MT':'Montana','NE':'Nebraska','NV':'Nevada','NH':'New Hampshire','NJ':'New Jersey',
 'NM':'New Mexico','NY':'New York','NC':'North Carolina','ND':'North Dakota','OH':'Ohio',
 'OK':'Oklahoma','OR':'Oregon','PA':'Pennsylvania','RI':'Rhode Island','SC':'South Carolina',
 'SD':'South Dakota','TN':'Tennessee','TX':'Texas','UT':'Utah','VT':'Vermont',
 'VA':'Virginia','WA':'Washington','WV':'West Virginia','WI':'Wisconsin','WY':'Wyoming',
}

# CompLevel -> (discipline, level id).  None = unmapped (flagged).
LEVEL_MAP = {
 'XS':('WAG','wag-silver'), 'XP':('WAG','wag-plat'), 'XD':('WAG','wag-diamond'),
 'L9':('WAG','wag-l9'),     'Open':('WAG','wag-open'),
 'Dev':('MAG','mag-dev'),   'Int':('MAG','mag-int'), 'Adv':('MAG','mag-adv'),
 'NC':('MAG','mag-adv'),    'L7':('MAG','mag-dev'),  # renamed: NC->Advanced, L7->Developmental
 'MSTR':('MAG','mag-masters'),
 'NF':('TNT','tnt-new'),    'IF':('TNT','tnt-int'),  'HF':('TNT','tnt-high'),
}
GENDER = {'M':'Male','F':'Female','O':'Other'}

def sql_str(v):
    if v is None: return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

# ---------------------------------------------------------------- clubs
raw_clubs = json.load(open(p('clubs_raw.json'), encoding='utf-8'))
clubs = []
club_ids = set()
club_flags = []
for c in raw_clubs:
    cid = str(c['clubId'])
    club_ids.add(cid)
    abbr = (c.get('stateAbbr') or '').strip()
    full_state = ABBR_FULL.get(abbr, abbr)            # intl codes pass through
    sf_region = (c.get('Region') or '').strip()
    if sf_region in ENUM_REGIONS:
        region = sf_region
    elif full_state in STATE_REGIONS:
        region = STATE_REGIONS[full_state]
        club_flags.append(f"{c['name']}: SF region '{sf_region}' -> derived '{region}' from {full_state}")
    else:
        region = 'Other'
        club_flags.append(f"{c['name']}: SF region '{sf_region}', state '{abbr}' -> 'Other'")
    short = (c.get('Short Name') or '').strip() or c['name']
    clubs.append({
        'id': cid, 'name': c['name'], 'shortName': short,
        'state': full_state, 'region': region,
        'email': (c.get('Club Email') or '').strip(),
        'allowClubPay': False,
        '_sfDisplayId': c.get('ID'), '_managers': c.get('managers', []),
    })

# ---------------------------------------------------------------- people
wb = openpyxl.load_workbook(p('athletes_raw.xlsx'), read_only=True)
ws = wb.active
H = {h:i for i,h in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}
people = []
seen_ids = set()
dup_ids = []
unmapped_levels = {}
missing_club = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    aid = str(r[H['AthleteID']])
    if aid in seen_ids:
        dup_ids.append(aid); continue
    seen_ids.add(aid)
    cid = str(r[H['ClubID']])
    if cid not in club_ids:
        missing_club[cid] = missing_club.get(cid, 0) + 1
        cid = None
    dob = r[H['DOB']]
    dob = dob.date().isoformat() if isinstance(dob, datetime.datetime) else (dob.isoformat() if isinstance(dob, datetime.date) else None)
    gy = r[H['GradYear']]
    try: gy = int(gy) if gy not in (None,'') else None
    except (ValueError, TypeError): gy = None
    comp = (r[H['CompLevel']] or '').strip()
    levels = {}
    if comp in LEVEL_MAP:
        disc, lid = LEVEL_MAP[comp]; levels[disc] = lid
    elif comp:
        unmapped_levels[comp] = unmapped_levels.get(comp, 0) + 1
    people.append({
        'id': aid, 'kind': 'athlete',
        'firstName': (r[H['FirstName']] or '').strip(),
        'lastName': (r[H['LastName']] or '').strip(),
        'email': (r[H['Email']] or '').strip(),
        'dob': dob,
        'gender': GENDER.get((r[H['Gender']] or '').strip(), 'Other'),
        'gradYear': gy,
        'studentStatus': 'Student' if r[H['Student']] else 'Non-Student',
        'mainClubId': cid,
        'levels': levels,
        '_compLevel': comp,
        '_masterAthleteId': r[H['MasterAthleteID']],
    })

# ---------------------------------------------------------------- write refs
json.dump(clubs, open(p('clubs.json'),'w',encoding='utf-8'), indent=1)
json.dump(people, open(p('people.json'),'w',encoding='utf-8'), indent=1)

# ---------------------------------------------------------------- SQL
lines = []
lines.append("-- ScoreFlippers -> UCG import (clubs + people).")
lines.append("-- Replaces demo/transactional data with real data. Run in Supabase SQL editor.")
lines.append("-- Keeps config: seasons, levels, regions, coupons, user_roles, auth.users.")
lines.append("begin;")
lines.append("")
lines.append("-- 1. Clear demo + transactional data (FK-safe order).")
for t in ['scores','registrations','squads','meet_sessions','meets','memberships',
          'person_alt_clubs','club_managers','cart_items','invoice_items','invoices',
          'club_requests','people','clubs']:
    lines.append(f"delete from {t};")
lines.append("")
lines.append("-- 2. Clubs (%d).%s" % (len(clubs), ""))
cols = "id, name, short_name, state, region, email, allow_club_pay"
lines.append(f"insert into clubs ({cols}) values")
vals = []
for c in clubs:
    vals.append(f"  ({sql_str(c['id'])}, {sql_str(c['name'])}, {sql_str(c['shortName'])}, "
                f"{sql_str(c['state'])}, {sql_str(c['region'])}, {sql_str(c['email'])}, false)")
lines.append(",\n".join(vals) + ";")
lines.append("")
lines.append("-- 3. People (%d)." % len(people))
pcols = "id, kind, first_name, last_name, email, dob, gender, grad_year, student_status, main_club_id, levels"
# batch in chunks of 500 for editor friendliness
B = 500
for i in range(0, len(people), B):
    chunk = people[i:i+B]
    lines.append(f"insert into {'people'} ({pcols}) values")
    vv = []
    for pp in chunk:
        dob = sql_str(pp['dob'])
        if dob != 'NULL': dob += '::date'
        gy = 'NULL' if pp['gradYear'] is None else str(pp['gradYear'])
        lv = sql_str(json.dumps(pp['levels'])) + '::jsonb'
        vv.append(f"  ({sql_str(pp['id'])}, 'athlete', {sql_str(pp['firstName'])}, "
                  f"{sql_str(pp['lastName'])}, {sql_str(pp['email'])}, {dob}, "
                  f"{sql_str(pp['gender'])}, {gy}, {sql_str(pp['studentStatus'])}, "
                  f"{sql_str(pp['mainClubId'])}, {lv})")
    lines.append(",\n".join(vv) + ";")
    lines.append("")
lines.append("commit;")
open(p('import.sql'),'w',encoding='utf-8').write("\n".join(lines))

# ---------------------------------------------------------------- report
rep = []
rep.append("# ScoreFlippers import report\n")
rep.append(f"- Clubs: **{len(clubs)}**")
rep.append(f"- People: **{len(people)}** (athletes)")
rep.append(f"- Duplicate AthleteIDs skipped: {len(dup_ids)} {dup_ids[:10]}")
rep.append(f"- People with ClubID not in club list (main_club_id set NULL): "
           f"{sum(missing_club.values())} across {len(missing_club)} ids {dict(list(missing_club.items())[:10])}")
rep.append(f"- Missing email people: {sum(1 for x in people if not x['email'])}")
rep.append(f"- Missing grad year: {sum(1 for x in people if x['gradYear'] is None)}")
rep.append("\n## Unmapped CompLevel codes (levels left empty, flag for Nate)")
for k,v in sorted(unmapped_levels.items(), key=lambda x:-x[1]):
    rep.append(f"- `{k}`: {v} athletes")
rep.append("\n## CompLevel -> level mapping used")
for k,(d,l) in LEVEL_MAP.items():
    rep.append(f"- `{k}` -> {d} / {l}")
rep.append("\n## Club region normalizations / fallbacks")
for f in club_flags:
    rep.append(f"- {f}")
open(p('REPORT.md'),'w',encoding='utf-8').write("\n".join(rep))

print("clubs:", len(clubs), "people:", len(people))
print("dup ids:", len(dup_ids), "missing-club refs:", sum(missing_club.values()), list(missing_club.items())[:5])
print("unmapped levels:", unmapped_levels)
print("import.sql bytes:", os.path.getsize(p('import.sql')))

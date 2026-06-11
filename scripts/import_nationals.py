"""
Import NAIGC Nationals 2026 finals-qualification results (WAG/MAG/TNT) into a
compact JSON test-data file for the UCG platform.

Reads:
  Nationals 2026/Nationals 2026 Results/Finals Qualification Outputs/{wag,mag,tnt}_results.xlsx

Writes:
  ucg-platform/public/data/nationals-2026.json
"""
import json
import os
import random
from datetime import datetime, timezone

import openpyxl

SRC_BASE = (
    r"C:\Users\nssha\Steinsharpe Dropbox\Nate Sharpe\Documents\Misc\Gymnastics\NAIGC"
    r"\Nationals 2026\Nationals 2026 Results\Finals Qualification Outputs"
)
OUT_PATH = (
    r"C:\Users\nssha\Steinsharpe Dropbox\Nate Sharpe\Documents\Misc\Gymnastics\NAIGC"
    r"\NAIGC Reg & Scoring Platform\ucg-platform\public\data\nationals-2026.json"
)

WAG_FILE = os.path.join(SRC_BASE, "wag_results.xlsx")
MAG_FILE = os.path.join(SRC_BASE, "mag_results.xlsx")
TNT_FILE = os.path.join(SRC_BASE, "tnt_results.xlsx")

# ---------------------------------------------------------------------------
# Level mapping
# ---------------------------------------------------------------------------

# Existing platform level ids (from src/lib/seed.ts)
EXISTING_LEVELS = {
    "wag-silver": {"id": "wag-silver", "discipline": "WAG", "name": "Xcel Silver", "svMax": 10.0, "vaults": 2, "order": 1},
    "wag-plat":   {"id": "wag-plat",   "discipline": "WAG", "name": "Xcel Platinum", "svMax": 10.0, "vaults": 2, "order": 2},
    "wag-diamond":{"id": "wag-diamond","discipline": "WAG", "name": "Xcel Diamond", "svMax": 10.0, "vaults": 2, "order": 3},
    "wag-l9":     {"id": "wag-l9",     "discipline": "WAG", "name": "Level 9", "svMax": 10.1, "vaults": 2, "order": 4},
    "wag-open":   {"id": "wag-open",   "discipline": "WAG", "name": "Open Scoring", "svMax": None, "vaults": 2, "order": 5},
    "tnt-new":    {"id": "tnt-new",    "discipline": "TNT", "name": "New Flyers", "svMax": None, "vaults": 1, "order": 1},
    "tnt-int":    {"id": "tnt-int",    "discipline": "TNT", "name": "Intermediate Flyers", "svMax": None, "vaults": 1, "order": 2},
    "tnt-high":   {"id": "tnt-high",   "discipline": "TNT", "name": "High Flyers", "svMax": None, "vaults": 1, "order": 3},
}

# Raw CompLevel (WAG) -> level id
WAG_LEVEL_MAP = {
    "Silver": "wag-silver",
    "Platinum": "wag-plat",
    "Diamond": "wag-diamond",
    "Level 9": "wag-l9",
    "NAIGC Open Scoring": "wag-open",
}

# Raw CompLevel (MAG) -> new level id (none of these exist on platform yet)
MAG_LEVEL_MAP = {
    "NAIGC Developmental": "mag-beg",
    "NAIGC Intermediate": "mag-int2",
    "NAIGC Advanced (GymACT)": "mag-adv2",
}

# Raw *_Level (TNT) -> level id
TNT_LEVEL_MAP = {
    "New Flyers": "tnt-new",
    "Intermediate Flyers": "tnt-int",
    "High Flyers": "tnt-high",
}

# New (non-platform) level definitions referenced above.
# order continues the viewer's LEVEL_ORDER (Open, XD, XP, XG, XS, L10, L9, then others)
# -- these MAG levels are "other" so order continues after the WAG ones (5).
NEW_LEVELS = {
    "mag-beg":  {"id": "mag-beg",  "discipline": "MAG", "name": "Developmental", "svMax": None, "vaults": 1, "order": 6},
    "mag-int2": {"id": "mag-int2", "discipline": "MAG", "name": "Intermediate", "svMax": None, "vaults": 1, "order": 7},
    "mag-adv2": {"id": "mag-adv2", "discipline": "MAG", "name": "Advanced (GymACT)", "svMax": None, "vaults": 1, "order": 8},
}

WAG_EVENTS = [("VT", "VT_Score", "Vault Place", "VT?"),
              ("UB", "UB_Score", "Bars Place", "UB?"),
              ("BB", "BM_Score", "Beam Place", "BB?"),
              ("FX", "FX_Score", "Floor Place", "FX?")]

MAG_EVENTS = [("FX", "FX_Score", "Floor Place", "FX?"),
              ("PH", "PH_Score", "PH Place", "PH?"),
              ("SR", "SR_Score", "Rings Place", "SR?"),
              ("VT", "VT_Score", "Vault Place", "VT?"),
              ("PB", "PB_Score", "PB Place", "PB?"),
              ("HB", "HB_Score", "HB Place", "HB?")]

TNT_EVENTS = [("TU", "Tumbling_Score", "Tumbling Place", "Tumbling?", "Tumbling_Level"),
              ("TR", "Trampoline_Score", "Trampoline Place", "Trampoline?", "Trampoline_Level"),
              ("DM", "DMT_Score", "DMT Place", "DMT?", "DMT_Level")]


def truthy(v):
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip().lower() in ("y", "yes", "true", "1")
    if isinstance(v, (int, float)):
        return bool(v)
    return bool(v)


def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def round3(v):
    return round(float(v), 3)


def gender_map(g):
    if g is None:
        return "Other"
    g = str(g).strip().upper()
    if g in ("F", "FEMALE"):
        return "Female"
    if g in ("M", "MALE"):
        return "Male"
    return "Other"


def short_name(name):
    """Best-effort club shortName: trim common boilerplate words."""
    n = name.strip()
    drop_prefixes = ["The "]
    for p in drop_prefixes:
        if n.startswith(p):
            n = n[len(p):]
    drop_phrases = [
        "University of ", "College of ",
    ]
    for p in drop_phrases:
        if n.startswith(p):
            n = n[len(p):]
    drop_suffixes = [
        " Gymnastics Club", " Gymnastics Team", " Club Gymnastics",
        " Gymnastics", " University", " College",
        " Adult Gymnastics Club",
    ]
    for s in drop_suffixes:
        if n.endswith(s):
            n = n[: -len(s)]
            break
    n = n.strip()
    return n if n else name.strip()


def read_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = []
    for r in it:
        if len(r) == 0:
            continue
        d = {}
        for i, col in enumerate(header):
            d[col] = r[i] if i < len(r) else None
        rows.append(d)
    wb.close()
    return rows


def get_or_create_club(clubs, club_index, name):
    if not name or not str(name).strip():
        return None
    name = str(name).strip()
    if name in club_index:
        return club_index[name]
    cid = f"nc{len(clubs) + 1}"
    club = {"id": cid, "name": name, "shortName": short_name(name)}
    clubs.append(club)
    club_index[name] = cid
    return cid


def get_or_create_person(people, person_index, first, last, email, gender, student, club_id):
    first = (str(first).strip() if first else "")
    last = (str(last).strip() if last else "")
    email = (str(email).strip().lower() if email else "")

    if email:
        key = email
    else:
        key = f"{first.lower()}|{last.lower()}"

    if key in person_index:
        pid = person_index[key]
        # backfill clubId if missing
        for p in people:
            if p["id"] == pid:
                if not p.get("clubId") and club_id:
                    p["clubId"] = club_id
                break
        return pid

    pid = f"na{len(people) + 1}"
    person = {
        "id": pid,
        "f": first,
        "l": last,
        "e": email,
        "g": gender_map(gender),
        "student": truthy(student),
    }
    if club_id:
        person["clubId"] = club_id
    people.append(person)
    person_index[key] = pid
    return pid


def main():
    clubs = []
    club_index = {}
    people = []
    person_index = {}
    regs = []

    levels_used = set()
    skipped = {"wag": 0, "mag": 0, "tnt": 0}

    # ---------------- WAG ----------------
    wag_rows = read_rows(WAG_FILE)
    aa_mismatches = 0
    aa_checked = []
    for row in wag_rows:
        first = row.get("FirstName")
        last = row.get("LastName")
        club_name = row.get("Club_Name")
        if (not first or not str(first).strip()) and (not last or not str(last).strip()) and (not club_name or not str(club_name).strip()):
            skipped["wag"] += 1
            continue

        events = {}
        scores = {}
        places = {}
        for code, score_col, place_col, flag_col in WAG_EVENTS:
            sv = row.get(score_col)
            if is_num(sv) and sv > 0:
                scores[code] = round3(sv)
                events[code] = True
                pl = row.get(place_col)
                if is_num(pl):
                    places[code] = int(pl) if float(pl).is_integer() else pl

        if not scores:
            skipped["wag"] += 1
            continue

        comp_level = row.get("CompLevel")
        level_id = WAG_LEVEL_MAP.get(comp_level)
        if level_id is None:
            raise ValueError(f"Unmapped WAG CompLevel: {comp_level!r}")
        levels_used.add(level_id)

        club_id = get_or_create_club(clubs, club_index, club_name)
        person_id = get_or_create_person(
            people, person_index, first, last, row.get("Athlete_Email"),
            row.get("Gender"), row.get("Student"), club_id,
        )

        quals = {}
        for code, _, _, flag_col in WAG_EVENTS:
            if code in events and truthy(row.get(flag_col)):
                quals[code] = True
        if truthy(row.get("Team?")):
            quals["Team"] = True

        reg = {
            "id": f"nr{len(regs) + 1}",
            "athleteId": person_id,
            "discipline": "WAG",
            "levelId": level_id,
            "category": row.get("Placement Category"),
            "events": list(scores.keys()),
            "scores": scores,
        }
        if club_id:
            reg["clubId"] = club_id
        if places:
            reg["places"] = places
        if quals:
            reg["quals"] = quals
        regs.append(reg)

        # AA sanity check sample
        aa_score = row.get("AA_Score")
        if is_num(aa_score):
            aa_checked.append((row, scores, aa_score))

    # AA sanity check on ~20 random WAG rows
    random.seed(42)
    sample = random.sample(aa_checked, min(20, len(aa_checked)))
    aa_diff_count = 0
    for row, scores, aa_score in sample:
        s = sum(scores.values())
        if abs(s - float(aa_score)) > 0.002:
            aa_diff_count += 1

    # ---------------- MAG ----------------
    mag_rows = read_rows(MAG_FILE)
    for row in mag_rows:
        first = row.get("FirstName")
        last = row.get("LastName")
        club_name = row.get("Club_Name")
        if (not first or not str(first).strip()) and (not last or not str(last).strip()) and (not club_name or not str(club_name).strip()):
            skipped["mag"] += 1
            continue

        events = {}
        scores = {}
        places = {}
        for code, score_col, place_col, flag_col in MAG_EVENTS:
            sv = row.get(score_col)
            if is_num(sv) and sv > 0:
                scores[code] = round3(sv)
                events[code] = True
                pl = row.get(place_col)
                if is_num(pl):
                    places[code] = int(pl) if float(pl).is_integer() else pl

        if not scores:
            skipped["mag"] += 1
            continue

        comp_level = row.get("CompLevel")
        level_id = MAG_LEVEL_MAP.get(comp_level)
        if level_id is None:
            raise ValueError(f"Unmapped MAG CompLevel: {comp_level!r}")
        levels_used.add(level_id)

        club_id = get_or_create_club(clubs, club_index, club_name)
        person_id = get_or_create_person(
            people, person_index, first, last, row.get("Athlete_Email"),
            row.get("Gender"), row.get("Student"), club_id,
        )

        quals = {}
        for code, _, _, flag_col in MAG_EVENTS:
            if code in events and truthy(row.get(flag_col)):
                quals[code] = True
        if truthy(row.get("Team?")):
            quals["Team"] = True

        reg = {
            "id": f"nr{len(regs) + 1}",
            "athleteId": person_id,
            "discipline": "MAG",
            "levelId": level_id,
            "category": row.get("Placement Category"),
            "events": list(scores.keys()),
            "scores": scores,
        }
        if club_id:
            reg["clubId"] = club_id
        if places:
            reg["places"] = places
        if quals:
            reg["quals"] = quals
        regs.append(reg)

    # ---------------- TNT ----------------
    tnt_rows = read_rows(TNT_FILE)
    for row in tnt_rows:
        first = row.get("FirstName")
        last = row.get("LastName")
        club_name = row.get("Club_Name")
        if (not first or not str(first).strip()) and (not last or not str(last).strip()) and (not club_name or not str(club_name).strip()):
            skipped["tnt"] += 1
            continue

        any_event = False
        club_id = get_or_create_club(clubs, club_index, club_name)
        person_id = get_or_create_person(
            people, person_index, first, last, row.get("Athlete_Email"),
            row.get("Gender"), row.get("Student"), club_id,
        )

        for code, score_col, place_col, flag_col, level_col in TNT_EVENTS:
            sv = row.get(score_col)
            if not (is_num(sv) and sv > 0):
                continue
            any_event = True

            level_raw = row.get(level_col)
            level_id = TNT_LEVEL_MAP.get(level_raw)
            if level_id is None:
                raise ValueError(f"Unmapped TNT level for {code}: {level_raw!r}")
            levels_used.add(level_id)

            scores = {code: round3(sv)}
            places = {}
            pl = row.get(place_col)
            if is_num(pl):
                places[code] = int(pl) if float(pl).is_integer() else pl

            quals = {}
            if truthy(row.get(flag_col)):
                quals[code] = True

            reg = {
                "id": f"nr{len(regs) + 1}",
                "athleteId": person_id,
                "discipline": "TNT",
                "levelId": level_id,
                "events": [code],
                "scores": scores,
            }
            if club_id:
                reg["clubId"] = club_id
            if places:
                reg["places"] = places
            if quals:
                reg["quals"] = quals
            regs.append(reg)

        if not any_event:
            skipped["tnt"] += 1
            # remove the person if they were only just created and have no other regs
            # (cheap cleanup: check if person_id used anywhere else)
            if not any(r["athleteId"] == person_id for r in regs):
                # remove from people / index
                people[:] = [p for p in people if p["id"] != person_id]
                keys_to_del = [k for k, v in person_index.items() if v == person_id]
                for k in keys_to_del:
                    del person_index[k]

    # ---------------- Levels ----------------
    levels = []
    for lid in levels_used:
        if lid in EXISTING_LEVELS:
            levels.append(dict(EXISTING_LEVELS[lid]))
        elif lid in NEW_LEVELS:
            levels.append(dict(NEW_LEVELS[lid]))
        else:
            raise ValueError(f"Unknown level id used: {lid}")

    # sort by discipline then order for readability
    disc_order = {"WAG": 0, "MAG": 1, "TNT": 2}
    levels.sort(key=lambda l: (disc_order.get(l["discipline"], 9), l["order"]))

    # ---------------- Assemble ----------------
    out = {
        "meta": {
            "name": "NAIGC Nationals 2026",
            "slug": "naigc-nationals-2026",
            "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "source": "Finals Qualification Outputs (prelims + qualifier flags)",
        },
        "levels": levels,
        "clubs": clubs,
        "people": people,
        "regs": regs,
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, separators=(",", ":"))

    # ---------------- Validation ----------------
    file_size_kb = os.path.getsize(OUT_PATH) / 1024.0

    with open(OUT_PATH, "r", encoding="utf-8") as f:
        reloaded = json.load(f)

    counts_by_disc = {"WAG": 0, "MAG": 0, "TNT": 0}
    total_scores = 0
    for r in reloaded["regs"]:
        counts_by_disc[r["discipline"]] += 1
        total_scores += len(r["scores"])

    print("=== VALIDATION REPORT ===")
    print(f"levels: {len(reloaded['levels'])}")
    print(f"clubs: {len(reloaded['clubs'])}")
    print(f"people: {len(reloaded['people'])}")
    print(f"regs: {len(reloaded['regs'])}")
    print(f"  WAG regs: {counts_by_disc['WAG']}")
    print(f"  MAG regs: {counts_by_disc['MAG']}")
    print(f"  TNT regs: {counts_by_disc['TNT']}")
    print(f"total scores across all regs: {total_scores}")
    print(f"file size: {file_size_kb:.1f} KB")
    print()
    print("Skipped rows (no name/club or zero scores):")
    print(f"  WAG: {skipped['wag']}")
    print(f"  MAG: {skipped['mag']}")
    print(f"  TNT: {skipped['tnt']}")
    print()
    print("Levels used:")
    for lv in levels:
        print(f"  {lv['id']}: {lv['discipline']} {lv['name']} (order={lv['order']}, svMax={lv['svMax']})")
    print()
    print(f"AA sanity check: {len(sample)} WAG rows sampled, {aa_diff_count} differ from AA_Score by > 0.002")
    print()
    print("Sample regs (one per discipline):")
    for disc in ["WAG", "MAG", "TNT"]:
        for r in reloaded["regs"]:
            if r["discipline"] == disc:
                print(f"  {disc}: {json.dumps(r)}")
                break


if __name__ == "__main__":
    main()
